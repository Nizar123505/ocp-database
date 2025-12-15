from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import glob
from django.conf import settings

from .models import ExcelFile, ExcelColumn, FileCache, SheetDataCache
from .serializers import (
    ExcelFileSerializer, 
    ExcelFileCreateSerializer, 
    ExcelColumnSerializer,
    UserSerializer
)

# Dossier contenant les fichiers Excel
EXCEL_FOLDER = settings.BASE_DIR.parent

# Dossier d'archive pour les fichiers supprimés
ARCHIVE_FOLDER = os.path.join(settings.BASE_DIR.parent, '_archives')

# Créer le dossier d'archive s'il n'existe pas
if not os.path.exists(ARCHIVE_FOLDER):
    os.makedirs(ARCHIVE_FOLDER)


def update_file_cache(filepath, filename=None, last_modified_by=None):
    """Mettre à jour le cache pour un fichier spécifique"""
    if filename is None:
        filename = os.path.basename(filepath)
    
    try:
        wb = load_workbook(filepath, read_only=True)
        
        # Compter le total des entrées et détails par feuille
        total_entries = 0
        sheets_info = []
        sheets_details = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Compter les colonnes
            columns_count = 0
            for cell in ws[1]:
                if cell.value:
                    columns_count += 1
            
            # Compter les lignes
            row_count = 0
            for row in ws.iter_rows(min_row=2, max_row=5000, values_only=True):
                if any(cell is not None for cell in row):
                    row_count += 1
            
            total_entries += row_count
            sheets_info.append(sheet_name)
            sheets_details[sheet_name] = {
                'columns': columns_count,
                'entries': row_count
            }
        
        file_stat = os.stat(filepath)
        
        # Préparer les données à mettre à jour
        defaults = {
            'name': filename.replace('.xlsx', ''),
            'file_path': filepath,
            'sheets_count': len(wb.sheetnames),
            'sheets_json': sheets_info,
            'sheets_details': sheets_details,
            'total_entries': total_entries,
            'file_size': file_stat.st_size,
            'file_modified': datetime.fromtimestamp(file_stat.st_mtime)
        }
        
        # Ajouter le dernier utilisateur qui a modifié si fourni
        if last_modified_by:
            defaults['last_modified_by'] = last_modified_by
        
        # Mettre à jour ou créer le cache
        cache, created = FileCache.objects.update_or_create(
            filename=filename,
            defaults=defaults
        )
        
        wb.close()
        return cache
    except Exception as e:
        print(f"Erreur cache pour {filename}: {e}")
        return None


def cache_sheet_data(file_cache, filepath, sheet_name):
    """Mettre en cache les données d'une feuille avec analyse des types de colonnes"""
    from datetime import time, timedelta
    
    def serialize_value(value):
        """Convertir les valeurs non-JSON en chaînes"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M")
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        if isinstance(value, timedelta):
            total_seconds = int(value.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return value
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        
        ws = wb[sheet_name]
        
        # Récupérer les en-têtes
        headers = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        
        for cell in first_row:
            if cell:
                col_name = str(cell).strip().replace('\n', ' ')
                headers.append(col_name)
        
        # Collecter toutes les valeurs par colonne pour l'analyse
        column_values = {header: [] for header in headers}
        
        # Récupérer les données et collecter les valeurs par colonne
        data = []
        raw_rows = []  # Stocker les lignes brutes pour l'analyse
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=1000, values_only=True), start=2):
            if any(cell is not None for cell in row):
                row_data = {"_row_id": row_idx}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        # Collecter la valeur brute pour l'analyse
                        column_values[header].append(value)
                        # Sérialiser pour le stockage
                        row_data[header] = serialize_value(value)
                data.append(row_data)
        
        wb.close()
        
        # Analyser chaque colonne pour déterminer son type réel
        columns_info = []
        for idx, header in enumerate(headers, start=1):
            values = column_values.get(header, [])
            
            # Analyser les données réelles de la colonne (en passant aussi le nom pour fallback)
            field_type, data_type = analyze_column_data_type(values, header)
            
            # Collecter des exemples de valeurs non-null
            sample_vals = [str(v)[:50] for v in values[:5] if v is not None and str(v).strip() != ''][:3]
            
            columns_info.append({
                "index": idx,
                "name": header,
                "field_type": field_type,
                "data_type": data_type,  # 'number', 'text', 'text_only', 'date', 'boolean', 'any'
                "required": is_required_field(header),
                "sample_values": sample_vals  # Exemples de valeurs
            })
        
        # Sauvegarder dans le cache
        sheet_cache, created = SheetDataCache.objects.update_or_create(
            file_cache=file_cache,
            sheet_name=sheet_name,
            defaults={
                'headers': headers,
                'columns_info': columns_info,
                'data': data,
                'rows_count': len(data)
            }
        )
        
        return sheet_cache
    except Exception as e:
        print(f"Erreur cache feuille {sheet_name}: {e}")
        import traceback
        traceback.print_exc()
        return None


def sync_all_files_cache():
    """Synchroniser le cache avec tous les fichiers du dossier"""
    pattern = os.path.join(EXCEL_FOLDER, '*.xlsx')
    files = glob.glob(pattern)
    
    existing_filenames = set()
    
    for filepath in files:
        filename = os.path.basename(filepath)
        if filename.startswith('~$'):
            continue
        existing_filenames.add(filename)
        
        # Vérifier si le fichier a été modifié depuis le dernier cache
        try:
            file_stat = os.stat(filepath)
            file_modified = datetime.fromtimestamp(file_stat.st_mtime)
            
            cache = FileCache.objects.filter(filename=filename).first()
            # Comparer les dates (ignorer le timezone pour la comparaison)
            needs_update = cache is None
            if cache and cache.file_modified:
                # Convertir en timestamp pour comparaison fiable
                cache_ts = cache.file_modified.timestamp() if hasattr(cache.file_modified, 'timestamp') else 0
                file_ts = file_modified.timestamp()
                needs_update = file_ts > cache_ts
            
            if needs_update:
                # Mettre à jour le cache du fichier
                file_cache = update_file_cache(filepath, filename)
                
                # Mettre à jour aussi le cache des données de chaque feuille
                if file_cache and file_cache.sheets_json:
                    for sheet_name in file_cache.sheets_json:
                        cache_sheet_data(file_cache, filepath, sheet_name)
        except Exception as e:
            print(f"Erreur sync cache pour {filename}: {e}")
    
    # NE PAS supprimer les fichiers du cache si aucun fichier physique n'est trouvé
    # Cela permet de garder les données importées via l'API même si les fichiers
    # physiques ne sont pas présents (ex: en production sur Render)
    if existing_filenames:
        # Seulement supprimer si on a trouvé au moins un fichier physique
        FileCache.objects.filter(is_deleted=False).exclude(filename__in=existing_filenames).update(is_deleted=True)
    
    # Supprimer les caches de feuilles orphelins
    SheetDataCache.objects.filter(file_cache__isnull=True).delete()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """Récupérer les informations de l'utilisateur connecté"""
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_excel_files(request):
    """Récupérer la liste des fichiers Excel disponibles (depuis le cache) - VERSION SIMPLE ET RAPIDE"""
    try:
        # PAS de synchronisation sur Render - juste récupérer depuis la base
        # Synchroniser seulement en local si des fichiers physiques existent
        is_render = os.environ.get('RENDER', False)
        if not is_render:
            try:
                if os.path.exists(EXCEL_FOLDER) and glob.glob(os.path.join(EXCEL_FOLDER, '*.xlsx')):
                    sync_all_files_cache()
            except:
                pass  # Ignorer les erreurs de sync
        
        # Récupérer depuis le cache - TRÈS RAPIDE
        cached_files = FileCache.objects.filter(is_deleted=False)
        
        excel_files = []
        for cache in cached_files:
            try:
                last_modified_by_info = None
                if cache.last_modified_by:
                    last_modified_by_info = {
                        "id": cache.last_modified_by.id,
                        "username": cache.last_modified_by.username,
                        "full_name": cache.last_modified_by.username
                    }
                
                excel_files.append({
                    "id": cache.filename.replace(' ', '_').replace('.xlsx', '') if cache.filename else str(cache.id),
                    "name": cache.name or "Sans nom",
                    "filename": cache.filename or "",
                    "path": getattr(cache, 'file_path', "") or "",
                    "sheets_count": cache.sheets_count or 0,
                    "sheets": cache.sheets_json or [],
                    "total_entries": cache.total_entries or 0,
                    "size": cache.file_size or 0,
                    "modified": cache.file_modified.isoformat() if cache.file_modified else "",
                    "last_modified_by": last_modified_by_info
                })
            except Exception as e:
                print(f"Erreur fichier {cache.id}: {e}")
                continue
        
        return Response({
            "files": excel_files,
            "total_files": len(excel_files)
        })
        
    except Exception as e:
        print(f"Erreur get_excel_files: {e}")
        return Response({"files": [], "total_files": 0, "error": str(e)})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def refresh_files_cache(request):
    """Rafraîchir manuellement le cache des fichiers"""
    try:
        sync_all_files_cache()
        count = FileCache.objects.count()
        return Response({
            "message": "Cache rafraîchi avec succès",
            "files_count": count
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_file_sheets(request, filename):
    """Récupérer les feuilles d'un fichier Excel spécifique (depuis le cache)"""
    try:
        import json as json_module
        from urllib.parse import unquote
        
        # Décoder le filename (peut contenir des %20, etc.)
        decoded_filename = unquote(filename)
        
        # Chercher par filename exact, ou par filename décodé, ou par nom
        cache = FileCache.objects.filter(filename=decoded_filename).first()
        if not cache:
            cache = FileCache.objects.filter(filename=filename).first()
        if not cache:
            # Essayer avec .xlsx ajouté
            cache = FileCache.objects.filter(filename=f"{decoded_filename}.xlsx").first()
        if not cache:
            # Chercher par nom
            cache = FileCache.objects.filter(name__icontains=decoded_filename.replace('.xlsx', '')).first()
        
        if cache:
            # Parser sheets_json si c'est une chaîne
            sheets_list = cache.sheets_json
            if isinstance(sheets_list, str):
                try:
                    sheets_list = json_module.loads(sheets_list)
                except:
                    sheets_list = []
            
            # Parser sheets_details si c'est une chaîne
            sheets_details = cache.sheets_details
            if isinstance(sheets_details, str):
                try:
                    sheets_details = json_module.loads(sheets_details)
                except:
                    sheets_details = {}
            
            sheets = []
            if sheets_list:
                for sheet_name in sheets_list:
                    details = sheets_details.get(sheet_name, {}) if sheets_details else {}
                    sheets.append({
                        "name": sheet_name,
                        "columns_count": details.get('columns', 0),
                        "entries_count": details.get('entries', 0)
                    })
            
            return Response({
                "filename": cache.filename,
                "file_name": cache.name,
                "sheets": sheets
            })
        
        return Response({"error": f"Fichier non trouvé: {filename}"}, status=404)
        
    except Exception as e:
        print(f"Erreur get_file_sheets: {e}")
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_sheet_columns(request, filename, sheet_name):
    """Récupérer les colonnes d'une feuille (depuis le cache)"""
    try:
        # Chercher dans le cache
        file_cache = FileCache.objects.filter(filename=filename).first()
        
        if file_cache:
            sheet_cache = SheetDataCache.objects.filter(
                file_cache=file_cache, 
                sheet_name=sheet_name
            ).first()
            
            if sheet_cache and sheet_cache.columns_info:
                return Response({
                    "filename": filename,
                    "sheet_name": sheet_name,
                    "columns": sheet_cache.columns_info
                })
            
            # Cache la feuille si pas encore fait
            filepath = os.path.join(EXCEL_FOLDER, filename)
            if os.path.exists(filepath):
                sheet_cache = cache_sheet_data(file_cache, filepath, sheet_name)
                if sheet_cache:
                    return Response({
                        "filename": filename,
                        "sheet_name": sheet_name,
                        "columns": sheet_cache.columns_info
                    })
        
        # Fallback: lire directement
        filepath = os.path.join(EXCEL_FOLDER, filename)
        if not os.path.exists(filepath):
            return Response({"error": "Fichier non trouvé"}, status=404)
        
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return Response({"error": f"Feuille '{sheet_name}' non trouvée"}, status=404)
        
        ws = wb[sheet_name]
        columns = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        
        for idx, cell in enumerate(first_row, start=1):
            if cell:
                col_name = str(cell).strip().replace('\n', ' ')
                field_type, data_type = guess_field_type(col_name)
                columns.append({
                    "index": idx,
                    "name": col_name,
                    "field_type": field_type,
                    "data_type": data_type,
                    "required": is_required_field(col_name)
                })
        
        wb.close()
        
        return Response({
            "filename": filename,
            "sheet_name": sheet_name,
            "columns": columns
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


def analyze_column_data_type(column_values, column_name=""):
    """
    Analyser les valeurs d'une colonne pour déterminer son type réel.
    Retourne: (field_type, data_type)
    - field_type: 'number', 'datetime-local', 'text', 'textarea', 'select-yesno'
    - data_type: 'number', 'date', 'text', 'text_only', 'boolean', 'any'
    """
    from datetime import time, timedelta
    import re
    
    if not column_values:
        # Pas de données, utiliser le nom de la colonne
        return guess_field_type(column_name)
    
    # Filtrer les valeurs non-nulles
    non_null_values = [v for v in column_values if v is not None and str(v).strip() != '']
    
    if not non_null_values:
        # Toutes les valeurs sont nulles, utiliser le nom de la colonne
        return guess_field_type(column_name)
    
    # Compteurs pour chaque type
    date_count = 0
    number_count = 0
    text_count = 0
    long_text_count = 0
    oui_non_count = 0
    total = len(non_null_values)
    
    for value in non_null_values:
        # Vérifier si c'est une date/datetime
        if isinstance(value, datetime):
            date_count += 1
            continue
        if isinstance(value, time):
            date_count += 1
            continue
        if isinstance(value, timedelta):
            number_count += 1  # Les durées sont considérées comme des nombres
            continue
            
        # Vérifier si c'est un nombre
        if isinstance(value, (int, float)):
            number_count += 1
            continue
        
        # Convertir en string pour analyse
        str_value = str(value).strip()
        
        # Vérifier Oui/Non
        if str_value.lower() in ['oui', 'non', 'yes', 'no', 'o', 'n']:
            oui_non_count += 1
            continue
        
        # Vérifier si c'est une date en format string
        date_patterns = [
            r'^\d{2}/\d{2}/\d{4}$',  # DD/MM/YYYY
            r'^\d{4}-\d{2}-\d{2}',   # YYYY-MM-DD
            r'^\d{2}-\d{2}-\d{4}$',  # DD-MM-YYYY
            r'^\d{2}\.\d{2}\.\d{4}$' # DD.MM.YYYY
        ]
        is_date_string = any(re.match(pattern, str_value) for pattern in date_patterns)
        if is_date_string:
            date_count += 1
            continue
        
        # Vérifier si c'est un nombre en format string
        # Accepte les nombres avec virgule ou point comme séparateur décimal
        number_pattern = r'^-?\d+([.,]\d+)?$'
        if re.match(number_pattern, str_value.replace(' ', '')):
            number_count += 1
            continue
        
        # C'est du texte
        if len(str_value) > 100:
            long_text_count += 1
        else:
            text_count += 1
    
    # Déterminer le type dominant (au moins 60% des valeurs)
    threshold = 0.6
    
    if total > 0 and date_count / total >= threshold:
        return 'datetime-local', 'date'
    
    if total > 0 and number_count / total >= threshold:
        return 'number', 'number'
    
    if total > 0 and oui_non_count / total >= threshold:
        return 'select-yesno', 'boolean'
    
    if total > 0 and long_text_count / total >= 0.4:
        return 'textarea', 'text'
    
    # Si les données sont du texte, vérifier si c'est du "text_only" 
    # (noms de navires, clients, etc. qui ne doivent pas être des nombres)
    name_lower = column_name.lower()
    text_only_keywords = [
        'navire', 'navires', 'ship', 'vessel',
        'client', 'customer', 
        'fournisseur', 'supplier', 'vendeur',
        'origine', 'origin', 'provenance',
        'destination', 'dest',
        'region', 'région',
        'port', 'quai', 'terminal',
        'agent', 'agents', 'transitaire', 'armateur',
        'surveillant', 'surveill',
        'qualité', 'qualite', 'quality',
        'type', 'catégorie', 'categorie',
        'incoterm', 'incoterme',
        'facturation',
        'famille'
    ]
    if any(word in name_lower for word in text_only_keywords):
        return 'text', 'text_only'
    
    # Par défaut c'est du texte libre
    return 'text', 'any'


def guess_field_type(column_name):
    """Deviner le type de champ en fonction du nom de la colonne (fallback quand pas de données)"""
    name_lower = column_name.lower().strip()
    
    # Numéro de ligne (N°, N, #)
    if name_lower in ['n°', 'n', '#', 'no', 'num', 'numero', 'numéro']:
        return 'number', 'number'
    
    # Dates - patterns étendus
    date_keywords = [
        'date', 'arrivée', 'arrivee', 'arriv', 
        'debut', 'début', 'fin', 
        'accostage', 'appareillage', 
        'nor', 'quai libre', 
        'pose passerelle', 'ordre',
        'connection', 'déconnection', 'deconnection',
        'draft', 'notice'
    ]
    if any(word in name_lower for word in date_keywords):
        return 'datetime-local', 'date'
    
    # Nombres - patterns étendus
    number_keywords = [
        'tonnage', 'tonne', 'poids', 'masse',
        'nombre', 'nbr', 'nb',
        '%', 'h2o', 'h2so4', 'p2o5', 'k2o',
        'fob', 'fret', 'cfr', 'pu', 'p.u',
        'cours', 'valeur', 'montant', 'prix', 'tarif', 'total',
        'loa', 'jour', 'jr',
        'cadence', 'performance', 'taux',
        'attente', 'séjour', 'sejour', 'durée', 'duree',
        'surestaries', 'surrestaries', 
        'temps', 'humidité', 'humidite', 'humidit',
        'quantité', 'quantite', 'qte',
        'volume', 'surface',
        'acconnage', 'assurance',
        'fwd', 'aft', 'trim',
        'fresh water', 'fw',
        'mouvements', 'mvt',
        'concentration', 'concent'
    ]
    if any(word in name_lower for word in number_keywords):
        return 'number', 'number'
    
    # Texte long - patterns étendus
    long_text_keywords = [
        'remarques', 'remarque',
        'commentaire', 'commentaires', 'comment',
        'description', 'desc',
        'evenements', 'événements', 'evenement',
        'observation', 'observations', 'observ',
        'cause', 'conflit',
        'etat', 'état'
    ]
    if any(word in name_lower for word in long_text_keywords):
        return 'textarea', 'text'
    
    # Sélection Oui/Non
    if 'oui/non' in name_lower or 'oui non' in name_lower:
        return 'select-yesno', 'boolean'
    
    # Texte obligatoire (noms propres qui ne doivent pas être des nombres)
    text_only_keywords = [
        'navire', 'navires', 'ship', 'vessel',
        'client', 'customer', 
        'fournisseur', 'supplier', 'vendeur',
        'origine', 'origin', 'provenance',
        'destination', 'dest',
        'region', 'région',
        'port de chargement',
        'agent', 'agents', 'transitaire', 'armateur',
        'surveillant', 'surveill',
        'qualité', 'qualite', 'quality',
        'type', 'catégorie', 'categorie', 'cat',
        'incoterm', 'incoterme',
        'facturation',
        'famille',
        'dum', 'ei', 'cde', 'n° ei', 'n° cde'
    ]
    if any(word in name_lower for word in text_only_keywords):
        return 'text', 'text_only'
    
    # Par défaut: texte libre (accepte tout)
    return 'text', 'any'


def is_required_field(column_name):
    """Déterminer si un champ est obligatoire"""
    name_lower = column_name.lower()
    required_fields = ['n°', 'n', 'navires', 'date b/l', 'date bl', 'tonnage']
    return any(field in name_lower for field in required_fields)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_sheet_data(request, filename, sheet_name):
    """Récupérer les données d'une feuille (depuis le cache - instantané)"""
    try:
        import json as json_module
        from urllib.parse import unquote
        
        # Décoder les paramètres URL
        decoded_filename = unquote(filename)
        decoded_sheet_name = unquote(sheet_name)
        
        # Chercher le fichier dans le cache
        file_cache = FileCache.objects.filter(filename=decoded_filename).first()
        if not file_cache:
            file_cache = FileCache.objects.filter(filename=filename).first()
        if not file_cache:
            file_cache = FileCache.objects.filter(filename=f"{decoded_filename}.xlsx").first()
        if not file_cache:
            file_cache = FileCache.objects.filter(name__icontains=decoded_filename.replace('.xlsx', '')).first()
        
        if file_cache:
            # Chercher la feuille
            sheet_cache = SheetDataCache.objects.filter(
                file_cache=file_cache, 
                sheet_name=decoded_sheet_name
            ).first()
            
            if not sheet_cache:
                sheet_cache = SheetDataCache.objects.filter(
                    file_cache=file_cache, 
                    sheet_name=sheet_name
                ).first()
            
            if sheet_cache:
                # Parser headers si c'est une chaîne JSON
                headers = sheet_cache.headers
                if isinstance(headers, str):
                    try:
                        headers = json_module.loads(headers)
                    except:
                        headers = []
                
                # Parser data si c'est une chaîne JSON
                data = sheet_cache.data
                if isinstance(data, str):
                    try:
                        data = json_module.loads(data)
                    except:
                        data = []
                
                return Response({
                    "filename": file_cache.filename,
                    "sheet_name": sheet_cache.sheet_name,
                    "headers": headers,
                    "data": data,
                    "total_rows": sheet_cache.rows_count or len(data)
                })
        
        return Response({"error": f"Données non trouvées pour {filename}/{sheet_name}"}, status=404)
        
    except Exception as e:
        print(f"Erreur get_sheet_data: {e}")
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_sheet_entry(request, filename, sheet_name):
    """Ajouter une nouvelle entrée dans une feuille"""
    try:
        import json as json_module
        from urllib.parse import unquote
        
        decoded_filename = unquote(filename)
        decoded_sheet_name = unquote(sheet_name)
        filepath = os.path.join(EXCEL_FOLDER, decoded_filename)
        
        # Mode 1: Fichier physique existe (développement local)
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            
            if decoded_sheet_name not in wb.sheetnames:
                wb.close()
                return Response({"error": f"Feuille '{decoded_sheet_name}' non trouvée"}, status=404)
            
            ws = wb[decoded_sheet_name]
            
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().replace('\n', ' '))
            
            next_row = ws.max_row + 1
            entry_data = request.data
            
            for col_idx, header in enumerate(headers, start=1):
                value = entry_data.get(header, "")
                if value == "" or value is None:
                    value = None
                ws.cell(row=next_row, column=col_idx, value=value)
            
            wb.save(filepath)
            wb.close()
            
            file_cache = update_file_cache(filepath, decoded_filename, last_modified_by=request.user)
            if file_cache:
                cache_sheet_data(file_cache, filepath, decoded_sheet_name)
            
            return Response({"message": "Entrée ajoutée avec succès", "row_number": next_row})
        
        # Mode 2: Pas de fichier physique (production Render) - sauvegarder en base
        else:
            file_cache = FileCache.objects.filter(filename=decoded_filename).first()
            if not file_cache:
                file_cache = FileCache.objects.filter(filename=filename).first()
            if not file_cache:
                file_cache = FileCache.objects.filter(name__icontains=decoded_filename.replace('.xlsx', '')).first()
            
            if not file_cache:
                return Response({"error": "Fichier non trouvé en base"}, status=404)
            
            sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=decoded_sheet_name).first()
            if not sheet_cache:
                sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=sheet_name).first()
            
            if not sheet_cache:
                return Response({"error": "Feuille non trouvée en base"}, status=404)
            
            # Charger les données existantes
            data = sheet_cache.data
            if isinstance(data, str):
                try:
                    data = json_module.loads(data)
                except:
                    data = []
            if not isinstance(data, list):
                data = []
            
            # Ajouter la nouvelle entrée
            new_entry = dict(request.data)
            new_entry['_row_id'] = len(data) + 2  # +2 car ligne 1 = headers
            data.append(new_entry)
            
            # Sauvegarder en base
            sheet_cache.data = data
            sheet_cache.rows_count = len(data)
            sheet_cache.save()
            
            # Mettre à jour le total du fichier
            file_cache.total_entries = sum(
                s.rows_count for s in SheetDataCache.objects.filter(file_cache=file_cache)
            )
            file_cache.last_modified_by = request.user
            file_cache.save()
            
            return Response({"message": "Entrée ajoutée avec succès (base de données)", "row_number": new_entry['_row_id']})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_sheet_entry(request, filename, sheet_name):
    """Modifier une entrée existante dans une feuille"""
    try:
        import json as json_module
        from urllib.parse import unquote
        
        row_id = request.data.get('_row_id')
        if not row_id:
            return Response({"error": "ID de ligne requis"}, status=400)
        
        decoded_filename = unquote(filename)
        decoded_sheet_name = unquote(sheet_name)
        filepath = os.path.join(EXCEL_FOLDER, decoded_filename)
        
        # Mode 1: Fichier physique existe
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            if decoded_sheet_name not in wb.sheetnames:
                wb.close()
                return Response({"error": f"Feuille non trouvée"}, status=404)
            
            ws = wb[decoded_sheet_name]
            headers = [str(c.value).strip() for c in ws[1] if c.value]
            
            for col_idx, header in enumerate(headers, start=1):
                if header in request.data and header != '_row_id':
                    value = request.data[header]
                    if value == "" or value is None:
                        value = None
                    ws.cell(row=int(row_id), column=col_idx, value=value)
            
            wb.save(filepath)
            wb.close()
            
            file_cache = update_file_cache(filepath, decoded_filename, last_modified_by=request.user)
            if file_cache:
                cache_sheet_data(file_cache, filepath, decoded_sheet_name)
            
            return Response({"message": "Entrée modifiée avec succès"})
        
        # Mode 2: Base de données (Render)
        else:
            file_cache = FileCache.objects.filter(filename=decoded_filename).first()
            if not file_cache:
                file_cache = FileCache.objects.filter(filename=filename).first()
            if not file_cache:
                return Response({"error": "Fichier non trouvé"}, status=404)
            
            sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=decoded_sheet_name).first()
            if not sheet_cache:
                sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=sheet_name).first()
            if not sheet_cache:
                return Response({"error": "Feuille non trouvée"}, status=404)
            
            data = sheet_cache.data
            if isinstance(data, str):
                try:
                    data = json_module.loads(data)
                except:
                    data = []
            
            # Trouver et modifier la ligne
            for i, row in enumerate(data):
                if row.get('_row_id') == row_id or row.get('_row_id') == int(row_id):
                    for key, value in request.data.items():
                        if key != '_row_id':
                            data[i][key] = value
                    break
            
            sheet_cache.data = data
            sheet_cache.save()
            
            file_cache.last_modified_by = request.user
            file_cache.save()
            
            return Response({"message": "Entrée modifiée avec succès (base de données)"})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_sheet_entry(request, filename, sheet_name):
    """Supprimer une entrée d'une feuille"""
    try:
        import json as json_module
        from urllib.parse import unquote
        
        row_id = request.query_params.get('row_id')
        if not row_id:
            return Response({"error": "ID de ligne requis"}, status=400)
        
        decoded_filename = unquote(filename)
        decoded_sheet_name = unquote(sheet_name)
        filepath = os.path.join(EXCEL_FOLDER, decoded_filename)
        
        # Mode 1: Fichier physique existe
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            if decoded_sheet_name not in wb.sheetnames:
                wb.close()
                return Response({"error": "Feuille non trouvée"}, status=404)
            
            ws = wb[decoded_sheet_name]
            ws.delete_rows(int(row_id))
            wb.save(filepath)
            wb.close()
            
            file_cache = update_file_cache(filepath, decoded_filename, last_modified_by=request.user)
            if file_cache:
                cache_sheet_data(file_cache, filepath, decoded_sheet_name)
            
            return Response({"message": "Entrée supprimée avec succès"})
        
        # Mode 2: Base de données (Render)
        else:
            file_cache = FileCache.objects.filter(filename=decoded_filename).first()
            if not file_cache:
                file_cache = FileCache.objects.filter(filename=filename).first()
            if not file_cache:
                return Response({"error": "Fichier non trouvé"}, status=404)
            
            sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=decoded_sheet_name).first()
            if not sheet_cache:
                sheet_cache = SheetDataCache.objects.filter(file_cache=file_cache, sheet_name=sheet_name).first()
            if not sheet_cache:
                return Response({"error": "Feuille non trouvée"}, status=404)
            
            data = sheet_cache.data
            if isinstance(data, str):
                try:
                    data = json_module.loads(data)
                except:
                    data = []
            
            # Supprimer la ligne par _row_id
            data = [row for row in data if row.get('_row_id') != int(row_id) and row.get('_row_id') != row_id]
            
            sheet_cache.data = data
            sheet_cache.rows_count = len(data)
            sheet_cache.save()
            
            file_cache.total_entries = sum(s.rows_count for s in SheetDataCache.objects.filter(file_cache=file_cache))
            file_cache.last_modified_by = request.user
            file_cache.save()
            
            return Response({"message": "Entrée supprimée avec succès (base de données)"})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_excel(request, filename):
    """Télécharger un fichier Excel"""
    try:
        filepath = os.path.join(EXCEL_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return Response({"error": "Fichier non trouvé"}, status=404)
        
        response = FileResponse(
            open(filepath, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_excel_file(request):
    """Importer un fichier Excel existant depuis l'ordinateur"""
    try:
        uploaded_file = request.FILES.get('file')
        
        if not uploaded_file:
            return Response({"error": "Aucun fichier fourni"}, status=400)
        
        # Vérifier l'extension
        filename = uploaded_file.name
        if not filename.endswith('.xlsx') and not filename.endswith('.xls'):
            return Response({"error": "Le fichier doit être au format Excel (.xlsx ou .xls)"}, status=400)
        
        # Convertir .xls en .xlsx si nécessaire
        if filename.endswith('.xls'):
            filename = filename[:-4] + '.xlsx'
        
        filepath = os.path.join(EXCEL_FOLDER, filename)
        
        # Si le fichier existe déjà, ajouter un suffixe
        base_name = filename[:-5]  # Remove .xlsx
        counter = 1
        while os.path.exists(filepath):
            filename = f"{base_name} ({counter}).xlsx"
            filepath = os.path.join(EXCEL_FOLDER, filename)
            counter += 1
        
        # Vérifier que c'est un fichier Excel valide
        try:
            wb = load_workbook(uploaded_file)
            sheets_info = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Compter les colonnes
                headers = [cell.value for cell in ws[1] if cell.value]
                # Compter les lignes
                row_count = 0
                for row in ws.iter_rows(min_row=2, max_row=5000, values_only=True):
                    if any(cell is not None for cell in row):
                        row_count += 1
                
                sheets_info.append({
                    "name": sheet_name,
                    "columns": len(headers),
                    "rows": row_count
                })
            
            # Sauvegarder le fichier
            wb.save(filepath)
            wb.close()
            
            # Mettre à jour le cache des métadonnées (avec l'utilisateur qui importe)
            file_cache = update_file_cache(filepath, filename, last_modified_by=request.user)
            
            # Mettre en cache TOUTES les données des feuilles
            if file_cache:
                for sheet_name in file_cache.sheets_json:
                    cache_sheet_data(file_cache, filepath, sheet_name)
            
            return Response({
                "message": "Fichier importé avec succès",
                "filename": filename,
                "sheets": sheets_info,
                "total_sheets": len(sheets_info)
            }, status=201)
            
        except Exception as e:
            return Response({"error": f"Le fichier n'est pas un fichier Excel valide: {str(e)}"}, status=400)
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_excel_file(request):
    """Créer un nouveau fichier Excel avec des colonnes personnalisées"""
    try:
        file_name = request.data.get('name', '').strip()
        columns = request.data.get('columns', [])
        
        if not file_name:
            return Response({"error": "Le nom du fichier est requis"}, status=400)
        
        if not columns or len(columns) == 0:
            return Response({"error": "Au moins une colonne est requise"}, status=400)
        
        # Nettoyer le nom du fichier
        file_name = file_name.replace('/', '-').replace('\\', '-')
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        
        filepath = os.path.join(EXCEL_FOLDER, file_name)
        
        # Vérifier si le fichier existe déjà
        if os.path.exists(filepath):
            return Response({"error": f"Un fichier avec ce nom existe déjà: {file_name}"}, status=400)
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Données"
        
        # Style pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ajouter les colonnes
        for col_idx, column in enumerate(columns, start=1):
            col_name = column.get('name', f'Colonne {col_idx}')
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Ajuster la largeur de la colonne
            ws.column_dimensions[cell.column_letter].width = max(15, len(col_name) + 5)
        
        # Sauvegarder le fichier
        wb.save(filepath)
        wb.close()
        
        # Mettre à jour le cache des métadonnées (avec l'utilisateur qui crée)
        file_cache = update_file_cache(filepath, file_name, last_modified_by=request.user)
        
        # Mettre en cache les données de la feuille
        if file_cache:
            cache_sheet_data(file_cache, filepath, "Données")
        
        return Response({
            "message": "Fichier créé avec succès",
            "filename": file_name,
            "columns_count": len(columns)
        }, status=201)
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_sheet_to_file(request, filename):
    """Ajouter une nouvelle feuille à un fichier existant"""
    try:
        filepath = os.path.join(EXCEL_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return Response({"error": "Fichier non trouvé"}, status=404)
        
        sheet_name = request.data.get('name', '').strip()
        columns = request.data.get('columns', [])
        
        if not sheet_name:
            return Response({"error": "Le nom de la feuille est requis"}, status=400)
        
        if not columns or len(columns) == 0:
            return Response({"error": "Au moins une colonne est requise"}, status=400)
        
        wb = load_workbook(filepath)
        
        # Vérifier si la feuille existe déjà
        if sheet_name in wb.sheetnames:
            wb.close()
            return Response({"error": f"Une feuille avec ce nom existe déjà: {sheet_name}"}, status=400)
        
        # Créer la nouvelle feuille
        ws = wb.create_sheet(title=sheet_name)
        
        # Style pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ajouter les colonnes
        for col_idx, column in enumerate(columns, start=1):
            col_name = column.get('name', f'Colonne {col_idx}')
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(15, len(col_name) + 5)
        
        wb.save(filepath)
        wb.close()
        
        return Response({
            "message": "Feuille créée avec succès",
            "sheet_name": sheet_name,
            "columns_count": len(columns)
        }, status=201)
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_excel_file(request, filename):
    """Archiver un fichier Excel (suppression douce - le fichier reste en base de données)"""
    try:
        from django.utils import timezone
        import shutil
        
        filepath = os.path.join(EXCEL_FOLDER, filename)
        
        # Vérifier si le fichier existe dans le cache
        file_cache = FileCache.objects.filter(filename=filename, is_deleted=False).first()
        
        if not file_cache and not os.path.exists(filepath):
            return Response({"error": "Fichier non trouvé"}, status=404)
        
        # Créer un nom unique pour l'archive (avec timestamp)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        archive_filename = f"{os.path.splitext(filename)[0]}_{timestamp}.xlsx"
        archive_path = os.path.join(ARCHIVE_FOLDER, archive_filename)
        
        # Déplacer le fichier vers l'archive (si le fichier physique existe)
        if os.path.exists(filepath):
            shutil.move(filepath, archive_path)
        
        # Mettre à jour ou créer le cache avec les informations de suppression
        if file_cache:
            file_cache.is_deleted = True
            file_cache.deleted_at = timezone.now()
            file_cache.deleted_by = request.user
            file_cache.archived_path = archive_path if os.path.exists(archive_path) else None
            file_cache.save()
        else:
            # Créer une entrée dans le cache pour le fichier archivé
            FileCache.objects.create(
                filename=filename,
                name=os.path.splitext(filename)[0],
                file_path=filepath,
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=request.user,
                archived_path=archive_path if os.path.exists(archive_path) else None,
                file_modified=timezone.now(),
                sheets_count=0,
                total_entries=0
            )
        
        return Response({
            "message": "Fichier archivé avec succès",
            "archived": True,
            "can_restore": True,
            "archived_at": timezone.now().isoformat()
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_archived_files(request):
    """Récupérer la liste des fichiers archivés (supprimés)"""
    try:
        archived_files = FileCache.objects.filter(is_deleted=True).order_by('-deleted_at')
        
        files_list = []
        for cache in archived_files:
            deleted_by_info = None
            if cache.deleted_by:
                user = cache.deleted_by
                deleted_by_info = {
                    "id": user.id,
                    "username": user.username,
                    "full_name": f"{user.first_name} {user.last_name}".strip() or user.username
                }
            
            files_list.append({
                "id": cache.id,
                "filename": cache.filename,
                "name": cache.name,
                "deleted_at": cache.deleted_at.isoformat() if cache.deleted_at else None,
                "deleted_by": deleted_by_info,
                "sheets_count": cache.sheets_count,
                "total_entries": cache.total_entries,
                "can_restore": cache.archived_path and os.path.exists(cache.archived_path)
            })
        
        return Response({
            "archived_files": files_list,
            "total": len(files_list)
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def restore_excel_file(request, file_id):
    """Restaurer un fichier archivé"""
    try:
        from django.utils import timezone
        import shutil
        
        file_cache = FileCache.objects.filter(id=file_id, is_deleted=True).first()
        
        if not file_cache:
            return Response({"error": "Fichier archivé non trouvé"}, status=404)
        
        # Vérifier si le fichier physique existe dans l'archive
        if not file_cache.archived_path or not os.path.exists(file_cache.archived_path):
            return Response({
                "error": "Le fichier physique n'existe plus dans l'archive",
                "data_preserved": True,
                "message": "Les données du fichier sont toujours dans la base de données"
            }, status=400)
        
        # Restaurer le fichier à son emplacement original
        original_path = os.path.join(EXCEL_FOLDER, file_cache.filename)
        
        # Si un fichier avec le même nom existe déjà, le renommer
        if os.path.exists(original_path):
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            base_name = os.path.splitext(file_cache.filename)[0]
            new_filename = f"{base_name}_restored_{timestamp}.xlsx"
            original_path = os.path.join(EXCEL_FOLDER, new_filename)
            file_cache.filename = new_filename
            file_cache.name = os.path.splitext(new_filename)[0]
        
        # Déplacer le fichier de l'archive vers le dossier principal
        shutil.move(file_cache.archived_path, original_path)
        
        # Mettre à jour le cache
        file_cache.is_deleted = False
        file_cache.deleted_at = None
        file_cache.deleted_by = None
        file_cache.archived_path = None
        file_cache.file_path = original_path
        file_cache.last_modified_by = request.user
        file_cache.file_modified = timezone.now()
        file_cache.save()
        
        return Response({
            "message": "Fichier restauré avec succès",
            "filename": file_cache.filename,
            "restored_at": timezone.now().isoformat()
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def permanent_delete_file(request, file_id):
    """Supprimer définitivement un fichier archivé (IRRÉVERSIBLE - réservé aux admins)"""
    try:
        # Vérifier si l'utilisateur est admin
        if not request.user.is_staff and not request.user.is_superuser:
            return Response({"error": "Seuls les administrateurs peuvent supprimer définitivement"}, status=403)
        
        file_cache = FileCache.objects.filter(id=file_id, is_deleted=True).first()
        
        if not file_cache:
            return Response({"error": "Fichier archivé non trouvé"}, status=404)
        
        # Supprimer le fichier physique de l'archive s'il existe
        if file_cache.archived_path and os.path.exists(file_cache.archived_path):
            os.remove(file_cache.archived_path)
        
        # NE PAS supprimer du cache - garder l'historique
        # Au lieu de supprimer, on marque comme supprimé définitivement
        file_cache.archived_path = None  # Le fichier physique n'existe plus
        file_cache.save()
        
        return Response({
            "message": "Fichier physique supprimé. Les métadonnées sont conservées dans la base de données.",
            "data_preserved": True
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ViewSet pour compatibilité avec l'ancien système
class ExcelFileViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les fichiers Excel (ancien système)"""
    queryset = ExcelFile.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ExcelFileCreateSerializer
        return ExcelFileSerializer
