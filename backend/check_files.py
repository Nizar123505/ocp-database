"""
Script pour verifier les fichiers Excel et leurs colonnes
"""
import os
import sys
import django

# Configuration Django
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from api.models import FileCache, SheetDataCache
from openpyxl import load_workbook

EXCEL_FOLDER = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def check_all_files():
    """Verifier tous les fichiers et leurs colonnes"""
    print("=" * 80)
    print("ANALYSE DETAILLEE DES FICHIERS EXCEL")
    print("=" * 80)
    
    # Verifier le dossier
    print(f"\nDossier Excel: {EXCEL_FOLDER}")
    
    # Lister les fichiers xlsx
    import glob
    pattern = os.path.join(EXCEL_FOLDER, '*.xlsx')
    files = glob.glob(pattern)
    
    print(f"Fichiers trouves: {len(files)}")
    for f in files:
        print(f"  - {os.path.basename(f)}")
    
    # Analyser chaque fichier
    for filepath in files:
        filename = os.path.basename(filepath)
        if filename.startswith('~$'):
            continue
            
        print("\n" + "=" * 80)
        print(f"FICHIER: {filename}")
        print("=" * 80)
        
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Recuperer les en-tetes (premiere ligne)
                headers = []
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                
                for cell in first_row:
                    if cell:
                        headers.append(str(cell).strip().replace('\n', ' '))
                
                # Compter les lignes de donnees
                row_count = 0
                for row in ws.iter_rows(min_row=2, max_row=5000, values_only=True):
                    if any(cell is not None for cell in row):
                        row_count += 1
                
                print(f"\n  FEUILLE: {sheet_name}")
                print(f"  Colonnes: {len(headers)} | Lignes: {row_count}")
                print(f"  En-tetes:")
                for i, h in enumerate(headers, 1):
                    print(f"    {i}. {h}")
            
            wb.close()
            
        except Exception as e:
            print(f"  ERREUR: {e}")
    
    print("\n" + "=" * 80)
    print("FIN DE L'ANALYSE")
    print("=" * 80)

if __name__ == '__main__':
    check_all_files()



