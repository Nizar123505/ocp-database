from openpyxl import load_workbook
import os

# Chemin du fichier Excel
file_path = os.path.join(os.path.dirname(__file__), '..', 'Base de donnée import  2025.xlsx')

print(f"Lecture du fichier: {file_path}")
print()

wb = load_workbook(file_path, read_only=True, data_only=True)

print('=== ANALYSE DU FICHIER EXCEL ===')
print(f'Nombre de feuilles: {len(wb.sheetnames)}')
print(f'Feuilles: {wb.sheetnames}')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'--- Feuille: {sheet_name} ---')
    
    # Récupérer les en-têtes (première ligne)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    print(f'Colonnes ({len(headers)}):')
    for i, h in enumerate(headers, 1):
        print(f'  {i}. {h}')
    
    # Compter les lignes de données
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            row_count += 1
    print(f'Lignes de données: {row_count}')
    print()

wb.close()







